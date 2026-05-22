"""Session-scoped dataset registry with DuckDB-backed query support."""

from __future__ import annotations

import re
import tempfile
from pathlib import Path
from typing import Any

from loguru import logger

from nanobot.agent.dataset.profile import DatasetProfile
from nanobot.config.schema import DatasetsConfig

_SELECT_RE = re.compile(r"^\s*SELECT\b", re.IGNORECASE)
_FORBIDDEN_SQL = re.compile(
    r"\b(INSERT|UPDATE|DELETE|DROP|CREATE|ALTER|COPY|PRAGMA|INSTALL|LOAD|"
    r"ATTACH|DETACH|EXPORT|IMPORT|VACUUM|CHECKPOINT)\b",
    re.IGNORECASE,
)


class DatasetRegistry:
    """Per-session registry of uploaded datasets.

    Holds an in-memory mapping from stable ``dataset_id`` to the file path
    and its profile.  The mapping is serialisable via :meth:`to_metadata` /
    :meth:`from_metadata` so that dataset availability survives session
    save/load cycles.

    Query execution is delegated to DuckDB (in-memory, no server process).
    """

    def __init__(self, config: DatasetsConfig) -> None:
        self._config = config
        self._datasets: dict[str, tuple[Path, DatasetProfile]] = {}
        self._counter: dict[str, int] = {}  # stem -> disambiguation counter

    # -- registration --------------------------------------------------------

    def register(self, path: Path, profile: DatasetProfile | None = None) -> str:
        """Register *path* as a dataset, generating a stable id.

        If *profile* is None, it is generated on the fly via
        :func:`~nanobot.agent.dataset.profile.generate_profile`.
        """
        if profile is None:
            from nanobot.agent.dataset.profile import generate_profile
            profile = generate_profile(path, max_sample_rows=self._config.max_profile_rows)

        dataset_id = self._make_id(path)
        profile.dataset_id = dataset_id
        self._datasets[dataset_id] = (path, profile)
        logger.info("Registered dataset {} ({})", dataset_id, path.name)
        return dataset_id

    def get(self, dataset_id: str) -> tuple[Path, DatasetProfile] | None:
        return self._datasets.get(dataset_id)

    def has(self, dataset_id: str) -> bool:
        return dataset_id in self._datasets

    def list_ids(self) -> list[str]:
        return list(self._datasets.keys())

    def list_summaries(self) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for did, (path, profile) in self._datasets.items():
            result.append({
                "dataset_id": did,
                "filename": path.name,
                "file_type": profile.file_type,
                "shape": list(profile.shape),
                "sheet_names": profile.sheet_names,
                "columns": [{"name": c.name, "dtype": c.dtype} for c in profile.columns],
            })
        return result

    # -- persistence ---------------------------------------------------------

    def to_metadata(self) -> dict[str, Any]:
        """Serialise state for ``session.metadata["datasets"]``."""
        out: dict[str, Any] = {}
        for did, (path, profile) in self._datasets.items():
            from dataclasses import asdict
            out[did] = {
                "path": str(path),
                "profile": asdict(profile),
            }
        return out

    def from_metadata(self, data: dict[str, Any]) -> None:
        """Rehydrate from ``session.metadata["datasets"]``."""
        for did, entry in data.items():
            if not isinstance(entry, dict):
                continue
            path_str = entry.get("path")
            if not path_str:
                continue
            path = Path(path_str)
            if not path.is_file():
                logger.warning("Dataset file no longer exists: {}", path)
                continue
            prof_data = entry.get("profile", {})
            profile = _profile_from_dict(prof_data)
            profile.dataset_id = did
            self._datasets[did] = (path, profile)

    # -- query ---------------------------------------------------------------

    def query(
        self,
        dataset_id: str,
        sql: str,
    ) -> str:
        """Execute a SELECT-only SQL query against *dataset_id*.

        Returns a formatted table string (markdown for small results, TSV
        for larger ones).  Results are limited by the configured row / char
        caps.
        """
        entry = self._datasets.get(dataset_id)
        if entry is None:
            return f"Error: dataset '{dataset_id}' not found. Use list_datasets to see available datasets."

        path, profile = entry
        sql = sql.strip()

        # Safety: SELECT-only
        if not _SELECT_RE.match(sql):
            return (
                "Error: only SELECT queries are allowed. "
                "Use standard SQL SELECT syntax (e.g. SELECT col1, col2 FROM data WHERE ...)"
            )
        if _FORBIDDEN_SQL.search(sql):
            return (
                "Error: query contains forbidden SQL keywords (INSERT/UPDATE/DELETE/DROP/CREATE/ALTER/COPY/etc). "
                "Only SELECT queries are supported."
            )

        # Append LIMIT if missing
        if "LIMIT" not in sql.upper():
            sql = f"{sql.rstrip(';')} LIMIT {self._config.max_query_rows}"

        try:
            import duckdb
        except ImportError:
            return "Error: duckdb is not installed. Dataset querying requires duckdb."

        ext = path.suffix.lower()

        try:
            conn = duckdb.connect()
            try:
                self._load_table(conn, path, ext, profile)
                result = conn.execute(sql).fetchall()
                col_names = [desc[0] for desc in conn.description]
            finally:
                conn.close()
        except Exception as e:
            logger.error("DuckDB query failed for {}: {}", dataset_id, e)
            return f"Error executing query: {e}"

        return _format_result(col_names, result, self._config.max_query_chars)

    @staticmethod
    def _load_table(
        conn: Any,
        path: Path,
        ext: str,
        profile: DatasetProfile,
    ) -> str:
        """Load *path* into DuckDB as table 'data', returning the table name."""
        abs_path = str(path.resolve())

        if ext in (".csv", ".tsv"):
            conn.execute(f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{abs_path}')")
            return "data"

        if ext == ".xlsx":
            # DuckDB cannot natively read XLSX.  Stream through openpyxl into
            # a temporary CSV and let DuckDB consume that.
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = profile.sheet_names[0] if profile.sheet_names else wb.sheetnames[0]
                ws = wb[sheet]
                with tempfile.NamedTemporaryFile(
                    mode="w", suffix=".csv", delete=False, encoding="utf-8"
                ) as tf:
                    import csv as csv_mod
                    writer = csv_mod.writer(tf)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(
                            str(v) if v is not None else "" for v in row
                        )
                    tmp = tf.name
                try:
                    conn.execute(f"CREATE TABLE data AS SELECT * FROM read_csv_auto('{tmp}')")
                finally:
                    Path(tmp).unlink(missing_ok=True)
            finally:
                wb.close()
            return "data"

        raise ValueError(f"Unsupported dataset type: {ext}")

    # -- helpers -------------------------------------------------------------

    def _make_id(self, path: Path) -> str:
        import re as _re
        stem = path.stem
        safe = _re.sub(r"[^a-zA-Z0-9_-]", "_", stem)[:40]
        idx = self._counter.get(safe, 0) + 1
        self._counter[safe] = idx
        return f"ds_{safe}_{idx}" if idx > 1 else f"ds_{safe}"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _profile_from_dict(data: dict[str, Any]) -> DatasetProfile:
    """Reconstruct a :class:`DatasetProfile` from a plain dict (metadata round-trip)."""
    from nanobot.agent.dataset.profile import ColumnProfile

    cols = []
    for c in data.get("columns", []):
        cols.append(ColumnProfile(
            name=c.get("name", ""),
            dtype=c.get("dtype", "string"),
            null_count=c.get("null_count", 0),
            unique_count=c.get("unique_count", 0),
            min_val=c.get("min_val"),
            max_val=c.get("max_val"),
            mean_val=c.get("mean_val"),
            sample_values=c.get("sample_values", []),
        ))
    return DatasetProfile(
        dataset_id=data.get("dataset_id", ""),
        source_path=data.get("source_path", ""),
        source_filename=data.get("source_filename", ""),
        file_type=data.get("file_type", ""),
        sheet_names=data.get("sheet_names", []),
        shape=tuple(data.get("shape", (0, 0))),
        columns=cols,
        sample_rows=data.get("sample_rows", []),
        file_size_bytes=data.get("file_size_bytes", 0),
        truncated=data.get("truncated", False),
        notes=data.get("notes", []),
    )


def _format_result(
    col_names: list[str],
    rows: list[tuple],
    max_chars: int,
) -> str:
    """Render a query result as text with truncation."""
    if not rows:
        return "(empty result — no rows returned)"

    # Build a compact TSV-like representation
    lines = ["\t".join(col_names)]
    for row in rows:
        lines.append("\t".join(str(v) for v in row))

    text = "\n".join(lines)
    truncated = False
    if len(text) > max_chars:
        text = text[:max_chars]
        truncated = True

    header = (
        f"Query result: {len(rows)} row(s), {len(col_names)} column(s)"
        + (" [TRUNCATED]" if truncated else "")
        + "\n\n"
    )
    return header + text
