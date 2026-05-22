"""Lightweight dataset profile generation for CSV, TSV, and XLSX files.

Profiles are produced via streaming reads — per-column statistics are
accumulated incrementally so that large files never materialise entirely
in memory.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from loguru import logger

_MAX_PROFILE_COLS = 50  # only compute numeric stats for the first N columns
_MAX_PROFILE_CHARS = 2500  # hard cap on profile text output


@dataclass
class ColumnProfile:
    name: str
    dtype: str  # "integer", "float", "string", "boolean", "mixed"
    null_count: int = 0
    unique_count: int = 0
    min_val: float | None = None
    max_val: float | None = None
    mean_val: float | None = None
    sample_values: list[str] = field(default_factory=list)


@dataclass
class DatasetProfile:
    dataset_id: str
    source_path: str
    source_filename: str
    file_type: str  # "csv", "tsv", "xlsx"
    sheet_names: list[str] = field(default_factory=list)
    shape: tuple[int, int] = (0, 0)  # (rows, cols)
    columns: list[ColumnProfile] = field(default_factory=list)
    sample_rows: list[list[str]] = field(default_factory=list)
    file_size_bytes: int = 0
    truncated: bool = False
    notes: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_profile(
    path: Path,
    dataset_id: str = "",
    *,
    max_sample_rows: int = 5,
) -> DatasetProfile:
    """Generate a :class:`DatasetProfile` from *path*.

    The file type is inferred from the extension.  Only the first sheet is
    profiled for XLSX workbooks (all sheet *names* are recorded).
    """
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return _profile_xlsx(path, dataset_id, max_sample_rows=max_sample_rows)
    if ext in (".csv", ".tsv"):
        delimiter = "\t" if ext == ".tsv" else ","
        return _profile_delimited(path, dataset_id, ext.lstrip("."), delimiter, max_sample_rows)
    raise ValueError(f"Unsupported dataset file type: {ext}")


def profile_to_llm_text(profile: DatasetProfile) -> str:
    """Render *profile* as a compact text block suitable for the LLM context.

    The output includes the dataset id, filename, shape, column summaries,
    and the first few sample rows.  It is capped at ``_MAX_PROFILE_CHARS``.
    """
    lines: list[str] = []
    lines.append(f"dataset_id: {profile.dataset_id}")
    lines.append(f"file: {profile.source_filename} ({profile.file_type})")
    lines.append(f"size: {profile.shape[0]} rows x {profile.shape[1]} columns")
    if profile.sheet_names and len(profile.sheet_names) > 1:
        lines.append(f"sheets: {', '.join(profile.sheet_names)}")
    if profile.truncated:
        lines.append("(profile was sampled — not all rows were scanned)")
    if profile.notes:
        for note in profile.notes:
            lines.append(f"[{note}]")

    lines.append("")
    lines.append("columns:")
    for col in profile.columns:
        extras: list[str] = []
        if col.dtype in ("integer", "float") and col.mean_val is not None:
            extras.append(f"mean={col.mean_val:.3g}")
        if col.min_val is not None:
            extras.append(f"min={col.min_val:.3g}")
        if col.max_val is not None:
            extras.append(f"max={col.max_val:.3g}")
        null_note = f", {col.null_count} nulls" if col.null_count else ""
        extra_str = f" ({', '.join(extras)})" if extras else ""
        samples = col.sample_values[:3]
        sample_str = ", ".join(repr(s) for s in samples)
        lines.append(f"  {col.name}: {col.dtype}{extra_str}{null_note}  e.g. {sample_str}")

    if profile.sample_rows:
        lines.append("")
        lines.append(f"sample rows (first {len(profile.sample_rows)}):")
        # header row
        header = [c.name for c in profile.columns]
        lines.append("  " + "\t".join(header))
        for row in profile.sample_rows:
            lines.append("  " + "\t".join(str(v) for v in row))

    text = "\n".join(lines)
    if len(text) > _MAX_PROFILE_CHARS:
        text = text[:_MAX_PROFILE_CHARS] + "\n... (profile truncated)"
    return text


# ---------------------------------------------------------------------------
# Delimited files (CSV / TSV)
# ---------------------------------------------------------------------------


def _profile_delimited(
    path: Path,
    dataset_id: str,
    file_type: str,
    delimiter: str,
    max_sample_rows: int,
) -> DatasetProfile:
    try:
        size = path.stat().st_size
    except OSError:
        size = 0

    try:
        with open(path, encoding="utf-8") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            try:
                headers = next(reader)
            except StopIteration:
                return DatasetProfile(
                    dataset_id=dataset_id or _default_id(path),
                    source_path=str(path),
                    source_filename=path.name,
                    file_type=file_type,
                    shape=(0, 0),
                    notes=["empty file — no header row"],
                )
            col_count = len(headers)
            stats = _init_stats(headers, col_count)
            sample_rows: list[list[str]] = []
            row_count = 0
            for row in reader:
                _update_stats(stats, row, col_count)
                if len(sample_rows) < max_sample_rows:
                    sample_rows.append(list(row))
                row_count += 1
    except UnicodeDecodeError:
        # Try latin-1 fallback
        with open(path, encoding="latin-1") as fh:
            reader = csv.reader(fh, delimiter=delimiter)
            try:
                headers = next(reader)
            except StopIteration:
                return DatasetProfile(
                    dataset_id=dataset_id or _default_id(path),
                    source_path=str(path),
                    source_filename=path.name,
                    file_type=file_type,
                    shape=(0, 0),
                    notes=["empty file"],
                )
            col_count = len(headers)
            stats = _init_stats(headers, col_count)
            sample_rows = []
            row_count = 0
            for row in reader:
                _update_stats(stats, row, col_count)
                if len(sample_rows) < max_sample_rows:
                    sample_rows.append(list(row))
                row_count += 1

    columns = _finalise_columns(stats)
    return DatasetProfile(
        dataset_id=dataset_id or _default_id(path),
        source_path=str(path),
        source_filename=path.name,
        file_type=file_type,
        shape=(row_count, col_count),
        columns=columns,
        sample_rows=sample_rows,
        file_size_bytes=size,
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _profile_xlsx(
    path: Path,
    dataset_id: str,
    max_sample_rows: int,
) -> DatasetProfile:
    try:
        size = path.stat().st_size
    except OSError:
        size = 0

    try:
        from openpyxl import load_workbook
    except ImportError:
        return DatasetProfile(
            dataset_id=dataset_id or _default_id(path),
            source_path=str(path),
            source_filename=path.name,
            file_type="xlsx",
            notes=["openpyxl not installed — cannot profile"],
        )

    notes: list[str] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("Failed to open XLSX {}: {}", path, e)
        return DatasetProfile(
            dataset_id=dataset_id or _default_id(path),
            source_path=str(path),
            source_filename=path.name,
            file_type="xlsx",
            notes=[f"failed to open: {e!s}"],
        )

    try:
        sheet_names = wb.sheetnames
        if not sheet_names:
            wb.close()
            return DatasetProfile(
                dataset_id=dataset_id or _default_id(path),
                source_path=str(path),
                source_filename=path.name,
                file_type="xlsx",
                shape=(0, 0),
                notes=["no sheets in workbook"],
            )

        target = sheet_names[0]
        ws = wb[target]
        rows_iter = ws.iter_rows(values_only=True)

        try:
            headers_raw = next(rows_iter)
        except StopIteration:
            wb.close()
            return DatasetProfile(
                dataset_id=dataset_id or _default_id(path),
                source_path=str(path),
                source_filename=path.name,
                file_type="xlsx",
                sheet_names=sheet_names,
                shape=(0, 0),
                notes=[f"sheet '{target}' is empty"],
            )

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(headers_raw)]
        col_count = len(headers)
        stats = _init_stats(headers, col_count)
        sample_rows: list[list[str]] = []
        row_count = 0

        for row in rows_iter:
            values = [str(v) if v is not None else "" for v in row]
            # Pad short rows
            if len(values) < col_count:
                values += [""] * (col_count - len(values))
            _update_stats(stats, values[:col_count], col_count)
            if len(sample_rows) < max_sample_rows:
                sample_rows.append(values[:col_count])
            row_count += 1

        # Add note about data_only mode
        notes.append("values are computed (data_only=True) — formulas are not shown")

    finally:
        wb.close()

    columns = _finalise_columns(stats)
    return DatasetProfile(
        dataset_id=dataset_id or _default_id(path),
        source_path=str(path),
        source_filename=path.name,
        file_type="xlsx",
        sheet_names=sheet_names,
        shape=(row_count, col_count),
        columns=columns,
        sample_rows=sample_rows,
        file_size_bytes=size,
        notes=notes,
    )


# ---------------------------------------------------------------------------
# Internal helpers — streaming statistics
# ---------------------------------------------------------------------------

_STAT_KEYS = ("count", "nulls", "sum", "min", "max", "uniques")


def _init_stats(headers: list[str], col_count: int) -> list[dict[str, Any]]:
    stats: list[dict[str, Any]] = []
    for i in range(min(col_count, _MAX_PROFILE_COLS)):
        stats.append({
            "name": headers[i] if i < len(headers) else f"col_{i}",
            "count": 0,
            "nulls": 0,
            "sum": 0.0,
            "min": None,
            "max": None,
            "uniques": set(),
            "type": None,  # "int", "float", "str", "bool", "mixed"
            "samples": [],
            "max_uniques": 200,
        })
    return stats


def _try_number(val: str) -> int | float | None:
    """Return int/float if *val* is a clean number, else None."""
    if not val or not val.strip():
        return None
    v = val.strip()
    try:
        if "." in v or "e" in v.lower():
            return float(v)
        return int(v)
    except ValueError:
        return None


def _update_stats(stats: list[dict[str, Any]], row: list[str], col_count: int) -> None:
    for i, st in enumerate(stats):
        if i >= col_count:
            break
        raw = row[i] if i < len(row) else ""
        val = raw.strip() if raw else ""

        if not val:
            st["nulls"] += 1
            st["count"] += 1
            continue

        st["count"] += 1

        # Samples
        if len(st["samples"]) < 3:
            st["samples"].append(val)

        # Uniques (bounded)
        uniq: set = st["uniques"]
        if len(uniq) < st["max_uniques"]:
            uniq.add(val)

        # Numeric stats
        num = _try_number(val)
        if num is not None:
            if st["type"] is None:
                st["type"] = "int" if isinstance(num, int) else "float"
            elif st["type"] == "int" and isinstance(num, float):
                st["type"] = "float"
            # else keep "str" / "mixed", type already downgraded

            st["sum"] += num
            if st["min"] is None or num < st["min"]:
                st["min"] = num
            if st["max"] is None or num > st["max"]:
                st["max"] = num
        else:
            if st["type"] in (None, "int", "float"):
                st["type"] = "str"
            elif st["type"] in ("int", "float") and st["count"] > 0:
                st["type"] = "mixed"


def _finalise_columns(stats: list[dict[str, Any]]) -> list[ColumnProfile]:
    columns: list[ColumnProfile] = []
    for st in stats:
        total = st["count"]
        dtype = st["type"] or "string"
        if dtype == "int":
            dtype = "integer"
        elif dtype == "str":
            dtype = "string"

        mean = (st["sum"] / (total - st["nulls"])) if (total - st["nulls"]) > 0 else None
        columns.append(ColumnProfile(
            name=st["name"],
            dtype=dtype,
            null_count=st["nulls"],
            unique_count=len(st["uniques"]),
            min_val=st["min"],
            max_val=st["max"],
            mean_val=mean,
            sample_values=st["samples"],
        ))
    return columns


def _default_id(path: Path) -> str:
    """Derive a stable dataset id from the filename."""
    import re
    stem = path.stem
    safe = re.sub(r"[^a-zA-Z0-9_-]", "_", stem)[:40]
    return f"ds_{safe}"
